import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

tables = [
    # ─── 기존 5개 테이블 ───
    {
        "table_name_en": "member",
        "table_name_kr": "회원",
        "description": "청첩장 플랫폼 사용자(신랑/신부) 계정 및 개인 정보 관리",
        "columns": [
            [1, "member_id", "회원ID", "VARCHAR", 50, "PK, NN", "회원 고유 식별자 (UUID 또는 아이디)", "", ""],
            [2, "password", "비밀번호", "VARCHAR", 255, "NN", "암호화된 비밀번호", "", ""],
            [3, "name", "이름", "VARCHAR", 50, "NN", "회원 이름", "", ""],
            [4, "email", "이메일", "VARCHAR", 100, "NN, UNIQUE", "이메일 주소", "", ""],
            [5, "phone", "전화번호", "VARCHAR", 20, "NN", "연락처", "", ""],
            [6, "subscription", "구독 상태", "VARCHAR", 20, "DEFAULT 'free'", "free / premium 구독 플랜 상태", "", "SaaS"],
            [7, "created_at", "가입일시", "TIMESTAMP", "", "NN", "계정 생성 일시", "", ""]
        ]
    },
    {
        "table_name_en": "invitation",
        "table_name_kr": "청첩장",
        "description": "회원이 생성한 모바일 청첩장 기본 정보",
        "columns": [
            [1, "inv_id", "청첩장ID", "VARCHAR", 50, "PK, NN", "청첩장 고유 식별자", "", ""],
            [2, "member_id", "회원ID", "VARCHAR", 50, "FK, NN", "청첩장을 생성한 회원", "member", ""],
            [3, "groom_name", "신랑이름", "VARCHAR", 50, "NN", "신랑 이름", "", ""],
            [4, "bride_name", "신부이름", "VARCHAR", 50, "NN", "신부 이름", "", ""],
            [5, "wedding_date", "예식일시", "DATETIME", "", "NN", "예식 날짜 및 시간", "", ""],
            [6, "location", "예식장소", "VARCHAR", 255, "NN", "예식장 주소", "", ""],
            [7, "bgm_url", "배경음악URL", "VARCHAR", 255, "", "업로드된 BGM 파일 경로", "", ""],
            [8, "created_at", "생성일시", "TIMESTAMP", "", "NN", "청첩장 생성 일시", "", ""]
        ]
    },
    {
        "table_name_en": "guestbook",
        "table_name_kr": "방명록",
        "description": "하객이 청첩장에 남긴 축하 메시지 관리",
        "columns": [
            [1, "gb_id", "방명록ID", "INT", "", "PK, NN, AI", "방명록 식별자 (Auto Increment)", "", ""],
            [2, "inv_id", "청첩장ID", "VARCHAR", 50, "FK, NN", "메시지가 작성된 청첩장", "invitation", ""],
            [3, "guest_name", "작성자명", "VARCHAR", 50, "NN", "하객 이름", "", ""],
            [4, "password", "비밀번호", "VARCHAR", 255, "NN", "글 삭제용 비밀번호", "", ""],
            [5, "message", "메시지내용", "TEXT", "", "NN", "축하 메시지", "", ""],
            [6, "created_at", "작성일시", "TIMESTAMP", "", "NN", "작성 일시", "", ""]
        ]
    },
    {
        "table_name_en": "rsvp",
        "table_name_kr": "참석여부",
        "description": "하객의 예식 참석 여부 및 동반인 수 관리",
        "columns": [
            [1, "rsvp_id", "RSVP_ID", "INT", "", "PK, NN, AI", "RSVP 식별자 (Auto Increment)", "", ""],
            [2, "inv_id", "청첩장ID", "VARCHAR", 50, "FK, NN", "대상 청첩장", "invitation", ""],
            [3, "guest_name", "참석자명", "VARCHAR", 50, "NN", "하객 이름", "", ""],
            [4, "contact", "연락처", "VARCHAR", 20, "", "하객 연락처", "", ""],
            [5, "attend_status", "참석여부", "VARCHAR", 10, "NN", "참석/불참/미정", "", ""],
            [6, "meal_status", "식사여부", "VARCHAR", 10, "", "식사/안함", "", ""],
            [7, "companion_count", "동반인수", "INT", "", "DEFAULT 0", "동반인 인원 수", "", ""],
            [8, "created_at", "제출일시", "TIMESTAMP", "", "NN", "제출 일시", "", ""]
        ]
    },
    {
        "table_name_en": "payment",
        "table_name_kr": "결제내역",
        "description": "간편 송금 연동을 통한 축의금 및 SaaS 구독 결제 내역",
        "columns": [
            [1, "pay_id", "결제ID", "VARCHAR", 100, "PK, NN", "결제 고유 식별자", "", ""],
            [2, "member_id", "회원ID", "VARCHAR", 50, "FK, NN", "결제한 회원", "member", ""],
            [3, "inv_id", "청첩장ID", "VARCHAR", 50, "FK", "대상 청첩장 (축의금인 경우)", "invitation", ""],
            [4, "pay_type", "결제유형", "VARCHAR", 20, "NN", "축의금 / SaaS구독", "", ""],
            [5, "sender_name", "송금자명", "VARCHAR", 50, "", "축의금 보낸 하객 이름", "", ""],
            [6, "amount", "금액", "INT", "", "NN", "송금 또는 구독 결제 금액", "", ""],
            [7, "status", "결제상태", "VARCHAR", 20, "NN", "완료/실패/취소", "", ""],
            [8, "created_at", "결제일시", "TIMESTAMP", "", "NN", "결제 일시", "", ""]
        ]
    },
    # ─── SaaS 추천 관련 신규 테이블 ───
    {
        "table_name_en": "preference",
        "table_name_kr": "웨딩 선호도",
        "description": "[SaaS] 회원이 입력한 웨딩 관련 선호도 정보 저장 (맞춤 추천 기반 데이터)",
        "columns": [
            [1, "pref_id", "선호도ID", "INT", "", "PK, NN, AI", "선호도 식별자 (Auto Increment)", "", "SaaS"],
            [2, "member_id", "회원ID", "VARCHAR", 50, "FK, NN", "선호도를 입력한 회원", "member", "SaaS"],
            [3, "budget_min", "최소예산", "INT", "", "NN", "희망 최소 예산 (만원 단위)", "", "SaaS"],
            [4, "budget_max", "최대예산", "INT", "", "NN", "희망 최대 예산 (만원 단위)", "", "SaaS"],
            [5, "region", "희망지역", "VARCHAR", 100, "NN", "희망 예식장 지역 (예: 서울, 경기)", "", "SaaS"],
            [6, "season", "희망계절", "VARCHAR", 20, "", "봄/여름/가을/겨울", "", "SaaS"],
            [7, "style", "선호스타일", "VARCHAR", 50, "", "모던/럭셔리/내추럴/클래식", "", "SaaS"],
            [8, "guest_count", "예상하객수", "INT", "", "", "예상 하객 인원 수", "", "SaaS"],
            [9, "created_at", "입력일시", "TIMESTAMP", "", "NN", "선호도 입력 일시", "", "SaaS"]
        ]
    },
    {
        "table_name_en": "recommendation",
        "table_name_kr": "추천 결과",
        "description": "[SaaS] 선호도 기반으로 시스템이 회원에게 추천한 웨딩 업체 정보 및 매칭 점수",
        "columns": [
            [1, "rec_id", "추천ID", "INT", "", "PK, NN, AI", "추천 결과 식별자 (Auto Increment)", "", "SaaS"],
            [2, "member_id", "회원ID", "VARCHAR", 50, "FK, NN", "추천 대상 회원", "member", "SaaS"],
            [3, "pref_id", "선호도ID", "INT", "", "FK, NN", "기반 선호도", "preference", "SaaS"],
            [4, "vendor_type", "업체유형", "VARCHAR", 30, "NN", "예식장 / 스튜디오 / 드레스 / 메이크업", "", "SaaS"],
            [5, "vendor_name", "업체명", "VARCHAR", 100, "NN", "추천된 업체 이름", "", "SaaS"],
            [6, "vendor_region", "업체지역", "VARCHAR", 100, "", "업체 소재 지역", "", "SaaS"],
            [7, "price_range", "가격대", "VARCHAR", 50, "", "예상 가격 범위 (만원)", "", "SaaS"],
            [8, "match_score", "매칭점수", "DECIMAL", "5,2", "NN", "선호도와의 매칭 점수 (0.00~100.00)", "", "SaaS"],
            [9, "created_at", "추천일시", "TIMESTAMP", "", "NN", "추천 생성 일시", "", "SaaS"]
        ]
    },
    {
        "table_name_en": "scrap",
        "table_name_kr": "스크랩(찜)",
        "description": "[SaaS] 회원이 추천받은 웨딩 업체 중 저장한 스크랩 목록",
        "columns": [
            [1, "scrap_id", "스크랩ID", "INT", "", "PK, NN, AI", "스크랩 식별자 (Auto Increment)", "", "SaaS"],
            [2, "member_id", "회원ID", "VARCHAR", 50, "FK, NN", "스크랩한 회원", "member", "SaaS"],
            [3, "rec_id", "추천ID", "INT", "", "FK, NN", "스크랩 대상 추천 결과", "recommendation", "SaaS"],
            [4, "memo", "메모", "TEXT", "", "", "회원이 추가한 개인 메모", "", "SaaS"],
            [5, "created_at", "스크랩일시", "TIMESTAMP", "", "NN", "스크랩 등록 일시", "", "SaaS"]
        ]
    }
]

header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
saas_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for idx, t in enumerate(tables):
    if idx == 0:
        ws = wb.active
        ws.title = t["table_name_en"]
    else:
        ws = wb.create_sheet(title=t["table_name_en"])

    ws['B1'] = "● 테이블 정의서"
    ws['B1'].font = Font(bold=True, size=14)

    ws['B2'] = "table name"
    ws['D2'] = t["table_name_en"]
    ws['G2'] = "테이블명"
    ws['H2'] = t["table_name_kr"]
    ws['B3'] = "설명"
    ws['D3'] = t["description"]

    headers = ["no", "column name", "컬럼명", "type", "length", "제약조건", "정의/설명", "참조테이블", "비고"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=5, column=i+2, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for r_idx, col in enumerate(t["columns"]):
        row_num = 6 + r_idx
        is_saas = (col[-1] == "SaaS")
        for c_idx, val in enumerate(col):
            cell = ws.cell(row=row_num, column=c_idx+2, value=val)
            cell.border = thin_border
            if c_idx not in [1, 2, 6]:
                cell.alignment = center_align
            if is_saas:
                cell.fill = saas_fill

    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 38
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 12

output_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\20260521_테이블정의서.xlsx'
wb.save(output_path)
print(f"Saved: {output_path}")
