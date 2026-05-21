import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "시스템 구성도"

# Define colors
blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
green_fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
orange_fill = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")
gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True, size=11)
black_font = Font(color="000000", bold=True, size=11)
title_font = Font(color="000000", bold=True, size=16)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Set row and col dimensions for grid
for col in range(1, 20):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

# Title
ws['B2'] = "● 시스템 구성도"
ws['B2'].font = title_font

def draw_box(ws, min_col, min_row, max_col, max_row, text, fill, font):
    # merge
    ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
    cell = ws.cell(row=min_row, column=min_col)
    cell.value = text
    cell.fill = fill
    cell.font = font
    cell.alignment = center_align
    # add border
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = thin_border

# Client Layer
draw_box(ws, 2, 5, 4, 7, "Client / User\n(Mobile/Desktop Browser)", blue_fill, white_font)

# Frontend Layer (Next.js)
draw_box(ws, 6, 5, 10, 7, "Frontend (Next.js 14)\n- App Router\n- Tailwind CSS", green_fill, white_font)

# Backend Layer (Supabase)
draw_box(ws, 12, 5, 16, 7, "Backend (Supabase)\n- Auth / PostgreSQL\n- Storage", orange_fill, white_font)

# External API Layer
draw_box(ws, 6, 10, 16, 11, "External APIs\n- Map (Naver/Kakao/Google)\n- Payment (PortOne)", gray_fill, black_font)

# Connections (Lines using arrows or text)
ws.merge_cells(start_row=6, start_column=5, end_row=6, end_column=5)
ws.cell(row=6, column=5, value="<-->").alignment = center_align

ws.merge_cells(start_row=6, start_column=11, end_row=6, end_column=11)
ws.cell(row=6, column=11, value="<-->").alignment = center_align

ws.merge_cells(start_row=8, start_column=8, end_row=9, end_column=8)
ws.cell(row=8, column=8, value="|").alignment = center_align

ws.merge_cells(start_row=8, start_column=14, end_row=9, end_column=14)
ws.cell(row=8, column=14, value="|").alignment = center_align

# Save the workbook
output_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\20260521_WeddingPlatform_시스템구성도.xlsx'
wb.save(output_path)
print(f"Saved to {output_path}")
