import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font

# Create new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "시스템 구성도"

# Title
ws['B2'] = "● 시스템 구성도"
ws['B2'].font = Font(color="000000", bold=True, size=16)

# Insert Image
img = Image(r'c:\Users\admin\Desktop\WeddingPlatform\temp\architecture.png')
img.anchor = 'B4'
ws.add_image(img)

# Save
output_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\20260521_시스템구성도.xlsx'
wb.save(output_path)
print(f"Saved to {output_path}")
