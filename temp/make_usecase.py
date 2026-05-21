import base64
import urllib.request
import json
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font

mermaid_code = """
flowchart LR
    %% Actors
    Guest(["👤 하객 (Guest)"])
    Host(["👤 회원/주최자 (Host)"])
    Admin(["👤 관리자 (Admin)"])

    %% Use Cases
    subgraph "Wedding Platform (모바일 청첩장 시스템)"
        direction TB
        UC1([청첩장 열람 및 3D 애니메이션 뷰])
        UC2([방명록 작성 및 본인글 삭제])
        UC3([참석 여부(RSVP) 제출])
        UC4([간편 축의금 송금])
        UC5([지도 길안내 확인])
        
        UC6([모바일 청첩장 생성 및 수정])
        UC7([BGM 오디오 파일 업로드])
        UC8([하객 응답 통계 및 방명록 조회])
        
        UC9([전체 청첩장 및 결제 관리])
    end

    %% Relationships
    Guest --> UC1
    Guest --> UC2
    Guest --> UC3
    Guest --> UC4
    Guest --> UC5

    Host --> UC6
    Host --> UC7
    Host --> UC8
    Host --> UC2

    Admin --> UC9
    
    style Guest fill:#f9f9f9,stroke:#333,stroke-width:2px
    style Host fill:#f9f9f9,stroke:#333,stroke-width:2px
    style Admin fill:#f9f9f9,stroke:#333,stroke-width:2px
"""

with open(r'c:\Users\admin\Desktop\WeddingPlatform\temp\usecase_diagram.mmd', 'w', encoding='utf-8') as f:
    f.write(mermaid_code.strip())

# Encode for mermaid.ink
state = {
    "code": mermaid_code.strip(),
    "mermaid": "{\n  \"theme\": \"default\"\n}",
    "autoSync": True,
    "updateDiagram": False
}
state_str = json.dumps(state)
b64 = base64.urlsafe_b64encode(state_str.encode('utf-8')).decode('ascii')
url = f"https://mermaid.ink/img/{b64}"

png_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\usecase_diagram.png'

try:
    print("Downloading Use Case Diagram PNG from mermaid.ink...")
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req) as response, open(png_path, 'wb') as out_file:
        out_file.write(response.read())
    print("Saved usecase_diagram.png successfully!")
except Exception as e:
    print(f"Error downloading PNG: {e}")

print("Creating Excel file...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "유스케이스 다이어그램"

ws['B2'] = "● 유스케이스 다이어그램"
ws['B2'].font = Font(color="000000", bold=True, size=16)

try:
    img = Image(png_path)
    img.anchor = 'B4'
    ws.add_image(img)
except Exception as e:
    print(f"Error adding image to excel: {e}")

output_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\20260521_유스케이스.xlsx'
wb.save(output_path)
print(f"Saved Excel successfully: {output_path}")
