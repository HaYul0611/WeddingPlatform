from plantuml import PlantUML
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font

puml_code = """
@startuml
!theme cerulean
skinparam defaultFontName Malgun Gothic
skinparam componentStyle rectangle

package "Client Layer" {
  [Mobile Web Browser] as Mobile
  [Desktop Web Browser] as Desktop
}

package "Frontend Layer (Next.js 14)" {
  [App Router (SSR/CSR)] as AppRouter
  [React Components] as Components
  [Tailwind CSS (Styling)] as Tailwind
}

package "Backend Layer (Supabase)" {
  [Supabase Auth] as Auth
  database "PostgreSQL DB" as DB
  [Supabase Storage] as Storage
}

cloud "External APIs" {
  [Map (Naver/Kakao/Google)] as MapAPI
  [Payment (PortOne)] as Payment
  [Notification (Alimtalk)] as Noti
}

Mobile --> AppRouter : HTTP / HTTPS
Desktop --> AppRouter : HTTP / HTTPS

AppRouter --> Auth : SDK API
AppRouter --> DB : SDK API
AppRouter --> Storage : SDK API

AppRouter --> MapAPI : REST
AppRouter --> Payment : REST
AppRouter --> Noti : REST

@enduml
"""

puml_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\architecture.puml'
with open(puml_path, 'w', encoding='utf-8') as f:
    f.write(puml_code)

print("Requesting image from PlantUML server...")
pl = PlantUML(url='http://www.plantuml.com/plantuml/img/')
png_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\architecture_puml.png'
pl.processes_file(puml_path, outfile=png_path)

print("Creating Excel file...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "시스템 구성도"

ws['B2'] = "● 시스템 구성도"
ws['B2'].font = Font(color="000000", bold=True, size=16)

img = Image(png_path)
img.anchor = 'B4'
ws.add_image(img)

output_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\20260521_시스템구성도.xlsx'
wb.save(output_path)
print("Saved Excel successfully using PlantUML.")
