import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "요구사항 일람"

# Title
ws['B1'] = "● 요구사항 일람"
ws['B1'].font = Font(bold=True, size=14)

header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

ws['B3'] = "번호"
ws['C3'] = "요구사항"
for col in ['B', 'C']:
    ws[f'{col}3'].font = header_font
    ws[f'{col}3'].fill = header_fill
    ws[f'{col}3'].alignment = center_align
    ws[f'{col}3'].border = thin_border

requirements = [
    # ─── 기존 청첩장 기능 ───
    "회원은 청첩장 플랫폼에 가입하기 위해 아이디, 비밀번호, 이름, 이메일, 전화번호를 입력해야 한다.",
    "회원은 가입 시 중복되지 않은 고유한 회원번호가 부여되며, 아이디는 중복될 수 없다.",
    "회원은 여러 개의 모바일 청첩장을 생성 및 관리할 수 있다.",
    "청첩장 생성 시 예식일, 예식장 위치, 신랑/신부 정보, 사진(갤러리), 인사말 등 기본 정보를 입력해야 한다.",
    "청첩장의 BGM은 사용자가 직접 오디오 파일을 업로드하여 설정할 수 있어야 한다.",
    "업로드된 BGM은 인스타그램 스타일의 몰입형 뮤직 플레이어 UI를 통해 재생되며 시각적인 구간 선택 기능을 제공해야 한다.",
    "청첩장 메인에는 예식일까지 남은 시간을 실시간으로 보여주는 D-Day 카운트다운(차감 방식) 기능이 표시되어야 한다.",
    "오시는 길 안내를 위해 네이버, 카카오, 구글 지도 연동 버튼이 동일한 라인에 SVG 아이콘 형태로 제공되어야 한다.",
    "하객은 청첩장에 접속하여 신랑/신부에게 축하 메시지를 남길 수 있는 방명록 기능을 이용할 수 있다.",
    "방명록에는 작성자 이름, 비밀번호, 내용이 포함되며, 비밀번호를 통해 본인이 작성한 글을 삭제할 수 있다.",
    "하객은 예식 참석 여부(RSVP) 및 식사 여부, 동반인 수를 청첩장을 통해 신랑/신부에게 전달할 수 있어야 한다.",
    "하객이 참석 여부를 제출하면, 신랑/신부 또는 하객에게 카카오 알림톡(또는 SMS)으로 알림이 발송되어야 한다.",
    "청첩장에는 신랑/신부에게 축의금을 보낼 수 있는 간편 송금 연동 기능이 포함되어야 한다.",
    "신랑/신부는 마이페이지를 통해 하객들의 참석 여부 통계 및 방명록 목록을 실시간으로 확인할 수 있어야 한다.",
    "관리자는 최고 관리자 대시보드를 통해 전체 회원 목록, 생성된 청첩장 수, 결제 및 서비스 이용 상태를 조회하고 관리할 수 있다.",
    # ─── SaaS 웨딩 맞춤형 추천 기능 ───
    "[SaaS] 회원은 예산, 희망 지역, 결혼 시즌(계절), 스타일(모던/럭셔리/내추럴 등) 선호도를 입력하여 맞춤 추천 설문을 제출할 수 있어야 한다.",
    "[SaaS] 시스템은 입력된 선호도를 기반으로 예식장, 스튜디오(스), 드레스(드), 메이크업(메) 등 웨딩 관련 업체를 자동으로 큐레이션하여 추천 결과를 제공해야 한다.",
    "[SaaS] 회원은 추천받은 웨딩 업체를 스크랩(찜)하여 저장하고, 저장된 항목끼리 비교 조회를 할 수 있어야 한다.",
    "[SaaS] 프리미엄 추천 기능은 구독 기반 SaaS 모델로 제공되며, 회원은 구독 플랜을 선택하여 결제 후 고도화된 맞춤형 큐레이션 서비스를 이용할 수 있어야 한다.",
]

for idx, req in enumerate(requirements, start=1):
    row = idx + 3
    ws.cell(row=row, column=2, value=idx)
    ws.cell(row=row, column=3, value=req)
    ws.cell(row=row, column=2).alignment = center_align
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=3).alignment = left_align
    ws.cell(row=row, column=3).border = thin_border

ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 120

output_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\20260521_요구사항일람.xlsx'
wb.save(output_path)
print(f"Saved: {output_path}")
