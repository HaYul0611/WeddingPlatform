import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "요구사항 정의서"

ws['B1'] = "● 요구사항 정의서"
ws['B1'].font = Font(bold=True, size=14)

headers = ["구분", "요구사항ID", "요구사항명", "기능ID", "기능명", "상세설명", "입력", "출력", "중요도", "기타"]
header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for col_idx, header_text in enumerate(headers, start=2):
    cell = ws.cell(row=2, column=col_idx, value=header_text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

legend = [
    "FR: 기능 요구사항 (Functional Requirements)",
    "NFR: 비기능 요구사항 (Non-Functional Requirements)",
    "TR: 기술 요구사항 (Technical Requirements)",
    "OR: 기타 요구사항 (Other Requirements)"
]
for idx, text in enumerate(legend):
    ws.cell(row=2+idx, column=13, value=text)

data = [
    # ─── 기존 청첩장 기능 ───
    ["FR", "FR-01", "회원 가입", "F-01", "회원 관리", "청첩장 플랫폼 가입을 위해 아이디, 비밀번호, 이름, 이메일, 전화번호 입력", "가입 폼 데이터", "가입 완료 알림", "상", "-"],
    ["FR", "FR-02", "회원 식별 및 계정 관리", "F-01", "회원 관리", "가입 시 중복되지 않은 고유 회원번호 부여 및 아이디 중복 확인", "아이디 조회", "고유번호 할당", "상", "-"],
    ["FR", "FR-03", "다중 청첩장 관리", "F-02", "청첩장 생성", "한 회원이 여러 개의 모바일 청첩장을 생성, 수정, 삭제 및 목록 조회 가능", "로그인 상태", "청첩장 목록 리스트", "상", "-"],
    ["FR", "FR-04", "청첩장 기본정보 등록", "F-02", "청첩장 생성", "예식일, 예식장 위치, 신랑/신부 정보, 사진(갤러리), 인사말 입력", "예식 정보 데이터", "청첩장 DB 저장", "상", "-"],
    ["FR", "FR-05", "BGM 파일 업로드", "F-02", "청첩장 생성", "사용자가 직접 원하는 BGM 오디오 파일을 업로드하여 배경음악 설정", "오디오 파일(mp3 등)", "업로드된 파일 URL", "중", "Storage"],
    ["FR", "FR-06", "뮤직 플레이어 UI", "F-03", "청첩장 뷰어", "업로드된 BGM을 인스타그램 스타일의 플레이어로 재생 및 시각적 구간 선택 기능 제공", "오디오 스트림", "오디오 재생 컨트롤", "중", "UI/UX"],
    ["FR", "FR-07", "D-Day 카운트다운 표시", "F-03", "청첩장 뷰어", "메인 화면에 예식일까지 남은 시간을 실시간 차감 방식으로 보여주는 위젯 제공", "예식일자", "차감된 남은 시간", "중", "UI/UX"],
    ["FR", "FR-08", "3사 지도 연동", "F-03", "청첩장 뷰어", "오시는 길 안내 시 네이버, 카카오, 구글 지도 연동 버튼을 동일 라인에 SVG 아이콘으로 배치", "위도/경도 좌표", "각 지도 앱/웹 링크 이동", "상", "외부연동"],
    ["FR", "FR-09", "하객 방명록 작성", "F-04", "하객 소통", "하객이 이름, 비밀번호, 축하 메시지를 남길 수 있는 방명록 기능", "방명록 폼", "등록된 메시지 노출", "상", "-"],
    ["FR", "FR-10", "방명록 본인 글 삭제", "F-04", "하객 소통", "방명록 작성 시 설정한 비밀번호를 통해 작성자 본인이 직접 글 삭제 기능", "글번호, 비밀번호", "삭제 완료", "중", "-"],
    ["FR", "FR-11", "RSVP (참석 여부) 제출", "F-04", "하객 소통", "하객이 예식 참석 여부, 식사 여부, 동반인 수를 폼을 통해 제출", "RSVP 폼 데이터", "응답 완료 알림", "상", "-"],
    ["FR", "FR-12", "알림톡 / SMS 자동 발송", "F-05", "알림 서비스", "RSVP 응답 완료 시 신랑/신부 또는 하객에게 카카오 알림톡이나 SMS로 알림 전송", "응답 완료 이벤트", "알림톡 메시지 발송", "상", "외부 API"],
    ["FR", "FR-13", "축의금 간편 송금 연동", "F-06", "결제 연동", "간편 송금 버튼 클릭 시 계좌 복사 또는 PortOne 간편 결제 시스템 연동 (모달 창)", "송금 요청", "결제/송금 완료", "상", "외부 API"],
    ["FR", "FR-14", "하객 통계 및 조회 대시보드", "F-07", "통계 조회", "신랑/신부가 마이페이지에서 하객들의 RSVP 참석 여부 통계 및 방명록 목록을 실시간 조회", "청첩장 ID", "통계 대시보드 화면", "중", "-"],
    ["FR", "FR-15", "최고 관리자 운영 대시보드", "F-08", "시스템 관리", "플랫폼 운영자가 전체 회원, 생성된 청첩장 수, 결제 상태 및 트래픽을 조회하고 관리", "관리자 권한", "어드민 페이지 뷰", "상", "-"],
    # ─── SaaS 웨딩 맞춤형 추천 기능 ───
    ["FR", "FR-16", "웨딩 선호도 입력", "F-09", "SaaS 추천", "회원이 예산, 희망 지역, 계절, 스타일(모던/럭셔리/내추럴) 등 선호도를 설문 폼으로 입력", "선호도 폼 데이터", "선호도 DB 저장", "상", "SaaS"],
    ["FR", "FR-17", "맞춤형 업체 큐레이션 조회", "F-09", "SaaS 추천", "입력된 선호도 기반으로 예식장·스튜디오·드레스·메이크업 업체를 매칭 점수와 함께 추천 결과로 제공", "선호도 프로필", "추천 업체 목록 + 매칭 점수", "상", "SaaS"],
    ["FR", "FR-18", "추천 업체 스크랩 및 비교", "F-09", "SaaS 추천", "회원이 추천받은 업체를 스크랩(찜)하여 저장하고, 저장 항목간 가격·특징 비교 가능", "스크랩 요청", "스크랩 목록 및 비교 화면", "중", "SaaS"],
    ["FR", "FR-19", "SaaS 구독 결제", "F-10", "SaaS 구독", "프리미엄 추천 기능 이용을 위해 구독 플랜(월/연) 선택 및 PortOne 결제 처리", "구독 플랜 선택 + 결제 정보", "구독 활성화 상태", "상", "SaaS/외부API"],
]

for row_idx, row_data in enumerate(data, start=3):
    for col_idx, cell_value in enumerate(row_data, start=2):
        cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
        cell.alignment = center_align if col_idx != 7 else left_align
        cell.border = thin_border

ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 65
ws.column_dimensions['H'].width = 22
ws.column_dimensions['I'].width = 22
ws.column_dimensions['J'].width = 8
ws.column_dimensions['K'].width = 12
ws.column_dimensions['L'].width = 5
ws.column_dimensions['M'].width = 48

output_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\20260521_요구사항정의서.xlsx'
wb.save(output_path)
print(f"Saved: {output_path}")
