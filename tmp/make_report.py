import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── 보고서 양식을 로드해서 SaaS 내용 추가 (기존 내용 + 추가) ──
template_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\8. 개인 프로젝트 결과 보고서.xlsx'
wb = openpyxl.load_workbook(template_path)
ws = wb.active

ws['C4'] = "WeddingPlatform (프리미엄 모바일 청첩장 + SaaS 웨딩 맞춤형 추천 플랫폼)"

ws['C8'] = (
    "RSVP·축의금 결제 등 통합 기능을 제공하는 차세대 인터랙티브 모바일 청첩장 서비스이자, "
    "회원의 예산·지역·스타일 선호도를 기반으로 예식장·스드메 업체를 자동 큐레이션하는 SaaS 형태의 웨딩 맞춤형 추천 플랫폼."
)

ws['C9'] = (
    "[ 청첩장 기능 ]\n"
    "- BGM 직접 업로드 및 뮤직 플레이어 UI 제공\n"
    "- D-Day 카운트다운(실시간 차감) 및 3사(네이버/카카오/구글) 지도 연동\n"
    "- 하객 방명록 작성/삭제 및 RSVP(참석/식사여부) 응답 제출\n"
    "- 카카오 알림톡 발송 및 PortOne 간편 축의금 송금 연동\n"
    "- 관리자 대시보드 및 하객 참석 통계 분석\n\n"
    "[ SaaS 웨딩 맞춤형 추천 기능 ]\n"
    "- 회원 선호도(예산/지역/계절/스타일) 설문 입력\n"
    "- 예식장·스튜디오·드레스·메이크업 업체 자동 큐레이션 및 매칭 점수 제공\n"
    "- 추천 업체 스크랩(찜) 및 항목간 가격·특징 비교 기능\n"
    "- 구독 플랜(월/연) 기반 프리미엄 SaaS 결제 모델"
)

ws['C10'] = (
    "- Next.js 서버 컴포넌트를 활용한 우수한 초기 로딩 속도 및 SEO 확보\n"
    "- Supabase 기반의 실시간 데이터 동기화 및 안전한 미디어 스토리지 구축\n"
    "- 다양한 외부 API (결제, 지도, 알림톡, 파트너 업체 DB) 모듈화 연동으로 확장성 극대화\n"
    "- 선호도 분석 알고리즘 기반의 초개인화 웨딩 큐레이션으로 결혼 준비 시간 단축\n"
    "- SaaS 구독 모델로 지속 가능한 수익 구조(Recurring Revenue) 실현"
)

ws['C11'] = (
    "- 개발환경 : Windows 11\n"
    "- 개발도구 : VS Code, Git\n"
    "- 언어 및 프레임워크 : TypeScript, Next.js 14 (App Router), Tailwind CSS\n"
    "- 백엔드 및 DB : Supabase (PostgreSQL, Auth, Storage, Realtime)\n"
    "- 외부 API : PortOne(결제·구독), Map APIs(Naver/Kakao/Google), 카카오 알림톡, 웨딩 파트너 API\n"
    "- AI 보조 도구 : Gemini (코드 생성, 아키텍처 설계, 설계 산출물 및 다이어그램 자동화 지원)\n"
    "- 서버 호스팅/배포 : Vercel"
)

output_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\20260521_개인프로젝트결과보고서.xlsx'
wb.save(output_path)
print(f"보고서 저장 완료: {output_path}")
