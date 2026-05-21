import openpyxl
import json

file_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\6. 유스케이스.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    output = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_data = []
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            sheet_data.append([str(cell) if cell is not None else "" for cell in row])
        output[sheet] = sheet_data
    
    with open(r'c:\Users\admin\Desktop\WeddingPlatform\temp\read_usecase.json', 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
except Exception as e:
    with open(r'c:\Users\admin\Desktop\WeddingPlatform\temp\read_usecase.json', 'w', encoding='utf-8') as f:
        f.write(str(e))
