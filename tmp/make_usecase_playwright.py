import asyncio
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from playwright.async_api import async_playwright

async def main():
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        page = await browser.new_page(viewport={'width': 1200, 'height': 800})
        await page.goto(r'file:///c:/Users/admin/Desktop/WeddingPlatform/temp/usecase.html')
        await page.wait_for_selector('svg') # wait for mermaid to render
        await asyncio.sleep(1) # Extra buffer
        element = page.locator('.mermaid')
        png_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\usecase_diagram.png'
        await element.screenshot(path=png_path)
        await browser.close()
        
        # Insert to Excel
        print("Creating Excel file...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "유스케이스 다이어그램"

        ws['B2'] = "● 유스케이스 다이어그램"
        ws['B2'].font = Font(color="000000", bold=True, size=16)

        img = Image(png_path)
        img.anchor = 'B4'
        ws.add_image(img)

        output_path = r'c:\Users\admin\Desktop\WeddingPlatform\temp\20260521_유스케이스.xlsx'
        wb.save(output_path)
        print(f"Saved Excel successfully: {output_path}")

asyncio.run(main())
